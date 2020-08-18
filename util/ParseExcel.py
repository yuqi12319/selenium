import time
import openpyxl
from openpyxl.styles import Border,Side,Font

class ParseExcel():
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)    #设置字体颜色
        #颜色对应的RGB值
        self.RGBDict = {'red':'FFFF3030','green':'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        #将Excel文件加载到内存，并获取其workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self,sheetName):
        #根据sheet名获取该sheet对象
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        #根据sheet的索引号获取该sheet对象
        try:
            sheet = self.workbook.worksheets[sheetIndex]
            return sheet
        except Exception as e:
            raise e

    def getRowsNumber(self,sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self,sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        #获取sheet中有数据区域的开始行号
        return sheet.min_row

    def getStartColNumber(self,sheet):
        # 获取sheet中有数据区域的开始列号
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple
        #下标从1开始，sheet.rows[1]表示第一行
        try:
            max_columns = sheet.max_column
            row_data = []
            for i in range(1,max_columns+1):
                cell_value = sheet.cell(row=rowNo,column=i).value
                row_data.append(cell_value)
            return row_data
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        # 获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple
        # 下标从1开始，sheet.columns[1]表示第一列
        try:
            max_rows = sheet.max_row
            column_data = []
            for i in range(1,max_rows+1):
                cell_value = sheet.cell(row=i,column=colNo).value
                column_data.append(cell_value)
            return column_data
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #根据单元格所在的位置索引获取该单元格中的值，下标从1开始
        #sheet.cell(row=1,column=1).value,表示Excel中第一行第一列的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordomates of cell!")

    def getCellOfObject(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #获取某个单元格的对象，可以根据单元格所在位置的数字索引,也可以直接根据Excel中单元格的编码及坐标
        #如getCellObject(sheet,coordinate='A1')or如getCellObject(sheet,rowNo=1,colsNo=2)
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordomates of cell!")

    # def writeCell(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        #根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        #下标从1开始，参数style表示字体的颜色的名字，比如red，green
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value = content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font = Font(color=self.RGBDict[style])
                print(self.excelFile)
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordomates of cell!")

    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #写入当前的时间，下标从1开始
        now = int(time.time())  #显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordomates of cell!")

if __name__ == '__main__':
    pe = ParseExcel()
    pe.loadWorkBook(u'G:\\pycharm\\python_workspace\\selenium\\testData\\126邮箱联系人.xlsx')
    print("通过名称获取sheet对象的名字:",pe.getSheetByName(u"联系人").title)
    print("通过index序号获取sheet对象的名字:",pe.getSheetByIndex(0).title)
    sheet = pe.getSheetByIndex(0)
    print(pe.getRowsNumber(sheet))  #获取最大行号
    print(pe.getColsNumber(sheet))  #获取最大列号
    rows = pe.getRow(sheet,2)       #获取某一行数据
    print(rows)
    columns = pe.getColumn(sheet,2) #获取某一列数据
    print(columns)
    #获取第一行第一列单元格内容
    print(pe.getCellOfValue(sheet,rowNo=1,colsNo=1))
    # pe.writeCell(sheet,u"我爱祖国",rowNo=10,colsNo=10)
    pe.writeCellCurrentTime(sheet,rowNo=10,colsNo=11)
